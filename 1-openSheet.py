import os
from openpyxl import load_workbook
import gspread
from oauth2client.service_account import ServiceAccountCredentials
from dotenv import load_dotenv  # Import dotenv to load environment variables

# Load environment variables from .env file
load_dotenv()

# Get the credentials.json file path from the .env file
CREDENTIALS_JSON = os.getenv('CREDENTIALS_JSON')  # Ensure this is set in your .env file
if not CREDENTIALS_JSON:
    raise ValueError("CREDENTIALS_JSON path not found in .env file.")

download_directory = os.path.join(os.getcwd(), "downloads")  
target_sheet_name = "CatalogFeedGoesHere" 
starting_column = "T"  

# Google Sheets setup
def setup_google_sheets():
    """
    Authenticate and connect to Google Sheets.
       """
    scope = [
        "https://spreadsheets.google.com/feeds",
        "https://www.googleapis.com/auth/spreadsheets",
        "https://www.googleapis.com/auth/drive.file",
        "https://www.googleapis.com/auth/drive"
    ]
    creds = ServiceAccountCredentials.from_json_keyfile_name(CREDENTIALS_JSON, scope)
    client = gspread.authorize(creds)
    return client

def append_data_to_google_sheet(download_directory, client, sheet_name, starting_column):
    """
    Append data from the downloaded Excel file starting at the specified column in the Google Sheet.
    """
    # Find the latest downloaded Excel file
    files = os.listdir(download_directory)
    excel_file = next((f for f in files if f.endswith(".xlsx")), None)
    if not excel_file:
        print("[ERROR] No Excel file found in the download directory.")
        return
    
    file_path = os.path.join(download_directory, excel_file)
    print(f"[INFO] Found Excel file: {file_path}")

    # Load the downloaded Excel file
    wb = load_workbook(file_path)
    sheet = wb.active

    # Extract data from the Excel sheet
    data = []
    for row in sheet.iter_rows(values_only=True):
        data.append(row)

    # Debug: Print extracted data
    print("[DEBUG] Data extracted from Excel:")
    for row in data:
        print(row)

    # Connect to Google Sheet and target the specified sheet tab
    gsheet = client.open("Admin1").worksheet(sheet_name)
    start_cell = f"{starting_column}3"

    # Update Google Sheet with data from Excel
    print("[INFO] Updating Google Sheet with actual data...")
    gsheet.update(start_cell, data)
    print(f"[INFO] Data successfully appended to the Google Sheet starting at {start_cell}.")

# Main execution block
try:
    client = setup_google_sheets()
    append_data_to_google_sheet(download_directory, client, target_sheet_name, starting_column)

except Exception as e:
    print(f"[ERROR] An error occurred: {e}")
